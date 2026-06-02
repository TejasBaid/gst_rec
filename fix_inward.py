"""
fix_inward.py
=============
Reads Book1.xlsx (Oracle ERP, 60 cols) + 2B file, reconciles, and rewrites
the Inward sheet of Final_April_2026 (6).xlsx with all 71 columns correct.

Columns 61-71 that we compute:
  61  State Code  = '07'  (always Delhi for this entity)
  62  State Name  = 'Delhi'
  63  GSTIN & INV = GSTIN(col21) + Supplier Invoice Num(col28)
  64  Total Tax   = SGST(col38) + CGST(col39) + IGST(col40)
  65  V-Lookuop   = 2B TOTAL TAX for matched invoice (per GSTIN+INV aggregate)
  66  Diff        = Oracle aggregate Total Tax - V-Lookuop
  67  Remark      = 'Matched' | 'RCM' | 'Not in 2B'
  68  TRX ID      = col15 (TRX ID repeated)
  69  ITC Type    = 'Fwd Rec'|'Fwd Non Rec'|'RCM Rec'|'RCM Non Rec'
  70  2B Year     = Inv Year tag (e.g. 'FY 2026-27')
  71  (count)     = count of Oracle lines per GSTIN+INV key for that row

ITC Type logic:
  TRX REC (col51) = 'Y' → recoverable,  'N' → non-recoverable
  RCM flag comes from 2B (Supply Attract Reverse Charge = 'Yes')
  → Y + non-RCM = 'Fwd Rec'
  → N + non-RCM = 'Fwd Non Rec'
  → Y + RCM     = 'RCM Rec'
  → N + RCM     = 'RCM Non Rec'
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import re, os

# ── paths ──────────────────────────────────────────────────────────────────────
ORACLE_FILE  = 'Book1.xlsx'
B2B_FILE     = 'DEL 07AABCG4768K1Z9_GSTR2B APR 26.xlsx'
FINAL_FILE   = 'Final_April_2026 (6).xlsx'
OUT_FILE     = 'Final_April_2026_CORRECTED.xlsx'

def clean(v):
    """Strip None / whitespace safely."""
    if v is None: return ''
    return str(v).strip()

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def make_key(gstin, inv):
    """Build GSTIN+INV lookup key (no normalisation — keep raw concat)."""
    return clean(gstin) + clean(inv)

# ── 1. Load Oracle (Book1.xlsx) ─────────────────────────────────────────────
print("Loading Oracle ERP …")
wb_ora = openpyxl.load_workbook(ORACLE_FILE, data_only=True)
ws_ora = wb_ora.worksheets[0]          # single sheet

# find header row
header_row_idx = None
for r_idx, row in enumerate(ws_ora.iter_rows(values_only=True), start=1):
    if row and any(str(v).upper().strip() == 'ACCOUNTING DATE' for v in row if v):
        header_row_idx = r_idx
        ora_headers = list(row)
        break
assert header_row_idx, "Oracle header not found"

# column index helpers (0-based inside the tuple)
def OC(name):
    """Return 0-based col index in Oracle row tuple."""
    for i, h in enumerate(ora_headers):
        if h and str(h).strip().upper() == name.upper():
            return i
    return None

OC_GSTIN   = OC('TP REGNNO')               # col 21
OC_INV     = OC('SUPPLIER INVOICE NUM')    # col 28
OC_SGST    = OC('SGST')                    # col 38
OC_CGST    = OC('CGST')                    # col 39
OC_IGST    = OC('IGST')                    # col 40
OC_TRX_REC = OC('TRX REC')                # col 51
OC_TRX_ID  = OC('TRX ID')                 # col 15

print(f"  Oracle cols → GSTIN={OC_GSTIN} INV={OC_INV} SGST={OC_SGST} CGST={OC_CGST} "
      f"IGST={OC_IGST} TRX_REC={OC_TRX_REC} TRX_ID={OC_TRX_ID}")

oracle_rows = []   # list of raw row tuples (values only)
for row in ws_ora.iter_rows(min_row=header_row_idx+1, values_only=True):
    if not any(v is not None for v in row): continue
    oracle_rows.append(row)

print(f"  Oracle data rows: {len(oracle_rows)}")

# ── 2. Load GSTR-2B (B2B+CN+DN sheet) ─────────────────────────────────────
print("Loading GSTR-2B …")
wb_2b = openpyxl.load_workbook(B2B_FILE, data_only=True)
ws_2b = wb_2b['B2B+CN+DN']

# header is row 3
b2b_headers = [cell.value for cell in list(ws_2b.iter_rows(min_row=3, max_row=3))[0]]

def B2C(name):
    for i, h in enumerate(b2b_headers):
        if h and str(h).strip() == name:
            return i
    return None

B2C_KEY   = B2C('GSTN & INV NO.')        # col 2 (0-idx 1)
B2C_GSTIN = B2C('GSTIN of supplier')     # col 3
B2C_INV   = B2C('Invoice number')        # col 5
B2C_RCM   = B2C('Supply Attract Reverse Charge')  # col 10
B2C_ITC   = B2C('ITC Availability')      # col 18
B2C_TOTAL = B2C('TOTAL TAX')             # col 25

print(f"  2B cols → KEY={B2C_KEY} GSTIN={B2C_GSTIN} INV={B2C_INV} "
      f"RCM={B2C_RCM} ITC={B2C_ITC} TOTAL={B2C_TOTAL}")

# Build per-key 2B lookup  { key → {total_tax, rcm, itc_avail} }
b2b_lookup = {}
for row in ws_2b.iter_rows(min_row=4, values_only=True):
    raw_key = row[B2C_KEY] if B2C_KEY is not None and B2C_KEY < len(row) else None
    if not raw_key: continue
    key = clean(raw_key)
    rcm = str(row[B2C_RCM]).strip().upper() == 'YES' if B2C_RCM is not None and B2C_RCM < len(row) else False
    itc = str(row[B2C_ITC]).strip() if B2C_ITC is not None and B2C_ITC < len(row) else 'Yes'
    total = num(row[B2C_TOTAL]) if B2C_TOTAL is not None and B2C_TOTAL < len(row) else 0.0
    b2b_lookup[key] = {'total_tax': total, 'rcm': rcm, 'itc_avail': itc}

print(f"  2B entries loaded: {len(b2b_lookup)}")

# ── 3. Compute aggregate Oracle total tax per GSTIN+INV key ──────────────────
# (so Diff = PR_aggregate - V-Lookuop)
agg_oracle_tax = defaultdict(float)
for row in oracle_rows:
    g = clean(row[OC_GSTIN] if OC_GSTIN is not None and OC_GSTIN < len(row) else '')
    inv = clean(row[OC_INV]  if OC_INV  is not None and OC_INV  < len(row) else '')
    key = make_key(g, inv)
    sgst = num(row[OC_SGST] if OC_SGST is not None and OC_SGST < len(row) else 0)
    cgst = num(row[OC_CGST] if OC_CGST is not None and OC_CGST < len(row) else 0)
    igst = num(row[OC_IGST] if OC_IGST is not None and OC_IGST < len(row) else 0)
    agg_oracle_tax[key] += sgst + cgst + igst

# ── 4. Build corrected Inward rows ───────────────────────────────────────────
print("Building corrected Inward rows …")

# Track count per key (for col 71)
key_line_count = defaultdict(int)
for row in oracle_rows:
    g   = clean(row[OC_GSTIN] if OC_GSTIN is not None and OC_GSTIN < len(row) else '')
    inv = clean(row[OC_INV]   if OC_INV   is not None and OC_INV   < len(row) else '')
    key_line_count[make_key(g, inv)] += 1

corrected_rows = []  # list of 71-element tuples
not_in_2b_count = 0
matched_count   = 0
rcm_count       = 0

for row in oracle_rows:
    row = list(row)    # make mutable

    # Pad to at least 60 elements
    while len(row) < 60:
        row.append(None)

    g   = clean(row[OC_GSTIN] if OC_GSTIN is not None else '')
    inv = clean(row[OC_INV]   if OC_INV   is not None else '')
    key = make_key(g, inv)

    sgst = num(row[OC_SGST] if OC_SGST is not None else 0)
    cgst = num(row[OC_CGST] if OC_CGST is not None else 0)
    igst = num(row[OC_IGST] if OC_IGST is not None else 0)
    total_tax_row = round(sgst + cgst + igst, 2)

    trx_rec  = clean(row[OC_TRX_REC] if OC_TRX_REC is not None else '').upper()
    trx_id   = row[OC_TRX_ID] if OC_TRX_ID is not None else None
    is_recov = (trx_rec == 'Y')

    b2b_entry = b2b_lookup.get(key)

    if b2b_entry:
        is_rcm   = b2b_entry['rcm']
        v_lookup = round(b2b_entry['total_tax'], 2)
        pr_agg   = round(agg_oracle_tax[key], 2)
        diff_val = round(pr_agg - v_lookup, 2)
        remark   = 'RCM' if is_rcm else 'Matched'

        if is_rcm:
            itc_type = 'RCM Rec' if is_recov else 'RCM Non Rec'
            rcm_count += 1
        else:
            itc_type = 'Fwd Rec' if is_recov else 'Fwd Non Rec'
            matched_count += 1
    else:
        # Not found in 2B — try alternative key variants
        alt_found = None
        if g and inv:
            # Try normalised: strip all non-alphanumeric
            norm_key = re.sub(r'[^A-Z0-9]', '', key.upper())
            for b_key in b2b_lookup:
                norm_b = re.sub(r'[^A-Z0-9]', '', b_key.upper())
                if norm_key == norm_b:
                    alt_found = b2b_lookup[b_key]
                    break

        if alt_found:
            is_rcm   = alt_found['rcm']
            v_lookup = round(alt_found['total_tax'], 2)
            pr_agg   = round(agg_oracle_tax[key], 2)
            diff_val = round(pr_agg - v_lookup, 2)
            remark   = 'RCM' if is_rcm else 'Matched'
            itc_type = ('RCM Rec' if is_recov else 'RCM Non Rec') if is_rcm else \
                       ('Fwd Rec' if is_recov else 'Fwd Non Rec')
            matched_count += 1
        else:
            v_lookup = None   # will show as #N/A equivalent
            pr_agg   = round(agg_oracle_tax[key], 2)
            diff_val = None
            remark   = 'Not in 2B'
            itc_type = 'Fwd Rec' if is_recov else 'Fwd Non Rec'
            not_in_2b_count += 1

    # Determine inv year
    inv_year = 'FY 2026-27'  # default for Apr-26 month
    # Simple heuristic: if invoice date < Apr 2026 → prior year
    # Oracle col 29 = SUPPLIER INVOICE DATE (0-indexed 28)
    inv_date_col = 28
    if inv_date_col < len(row) and row[inv_date_col]:
        d = row[inv_date_col]
        try:
            from datetime import datetime
            if hasattr(d, 'year'):
                if d.year < 2026 or (d.year == 2026 and d.month < 4):
                    inv_year = 'FY 2025-26'
        except: pass

    # Count of lines per this key
    line_count = key_line_count.get(key, 1)

    # Build 71-element row
    # cols 1-60 = Oracle data
    # cols 61-71 = computed
    full_row = list(row[:60])                   # Oracle cols 1-60
    full_row.append('07')                        # col 61 State Code
    full_row.append('Delhi')                     # col 62 State Name
    full_row.append(key if key else None)        # col 63 GSTIN & INV
    full_row.append(total_tax_row)               # col 64 Total Tax
    full_row.append(v_lookup)                    # col 65 V-Lookuop
    full_row.append(diff_val)                    # col 66 Diff
    full_row.append(remark)                      # col 67 Remark
    full_row.append(trx_id)                      # col 68 TRX ID (repeated)
    full_row.append(itc_type)                    # col 69 ITC Type
    full_row.append(inv_year)                    # col 70 2B Year
    full_row.append(line_count)                  # col 71 count

    corrected_rows.append(full_row)

print(f"  Rows built: {len(corrected_rows)}")
print(f"  Matched: {matched_count}  RCM: {rcm_count}  Not-in-2B: {not_in_2b_count}")

# ── 5. Verify ITC type totals match benchmark ─────────────────────────────────
itc_totals = defaultdict(lambda: {'igst': 0, 'cgst': 0, 'sgst': 0})
for row in corrected_rows:
    itc = row[68]   # col 69 ITC Type (0-indexed 68)
    sg  = num(row[37]) if 37 < len(row) else 0  # SGST col 38
    cg  = num(row[38]) if 38 < len(row) else 0  # CGST col 39
    ig  = num(row[39]) if 39 < len(row) else 0  # IGST col 40
    itc_totals[itc]['sgst'] += sg
    itc_totals[itc]['cgst'] += cg
    itc_totals[itc]['igst'] += ig

print("\n=== OUR CORRECTED ITC TYPE TOTALS ===")
for t, v in sorted(itc_totals.items()):
    print(f"  {t:20s}: IGST={v['igst']:>12,.2f}  CGST={v['cgst']:>12,.2f}  SGST={v['sgst']:>12,.2f}")

# Benchmark targets (from earlier analysis):
print("\n=== BENCHMARK ITC TYPE TARGETS ===")
bench = {
    'Fwd Non Rec':  {'igst': 101580.48, 'cgst': 116074.59, 'sgst': 116074.59},
    'Fwd Rec':      {'igst': 63923.60,  'cgst': 1044627.77,'sgst': 1044627.77},
    'RCM Non Rec':  {'igst': 12500.10,  'cgst': 18000.00,  'sgst': 18000.00},
    'RCM Rec':      {'igst': 8100.00,   'cgst': 147097.08, 'sgst': 147097.08},
}
for t, v in sorted(bench.items()):
    o = itc_totals.get(t, {'igst':0,'cgst':0,'sgst':0})
    di = round(v['igst'] - o['igst'], 2)
    dc = round(v['cgst'] - o['cgst'], 2)
    print(f"  {t:20s}: IGST_DIFF={di:>10,.2f}  CGST_DIFF={dc:>10,.2f}")

print("\nCorrected rows ready. Run fix_inward_write.py next to patch the Excel file.")
print(f"Saving to: {OUT_FILE}")

# Store corrected_rows for use in next step
import pickle
with open('/tmp/corrected_inward_rows.pkl', 'wb') as f:
    pickle.dump(corrected_rows, f)
print("Rows saved to /tmp/corrected_inward_rows.pkl")
EOF
